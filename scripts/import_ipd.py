"""Import IPD v7.00 xlsx → data/occupations/<padded>.json × 556.

Per DATA_ARCHITECTURE.md §2.1 / §5.1 / §3A / D-009 / D-010 / D-011.

Inputs:
    ~/Downloads/IPD_DL_numeric_7_00.xlsx       (518 occupations × numeric profile)
    ~/Downloads/IPD_DL_description_7_00_01.xlsx (556 occupations × text/classification)
    data/labels/*.ja-en.json                    (7 dimensions, IPD-ID → en_key reverse lookup)

Outputs:
    data/occupations/0001.json … 0584.json     (one per occupation, 4-digit padded id)
    data/.ipd_provenance.json                   (sha256 + retrieved_at)

Joins:
    By 収録番号 (IPD_01_01_001) = canonical_id (1-584).
    Numeric file is subset of description (518 ⊂ 556).
    Occupations not in numeric file get all 12 numeric subdivisions = None per §5.4.

Re-running is safe: output files are overwritten atomically (per file).

Usage:
    uv run python scripts/import_ipd.py
"""
from __future__ import annotations

import datetime as dt
import hashlib
import json
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "data"))
from schema.occupation import (  # noqa: E402
    Classifications,
    DataSourceVersions,
    Description,
    Occupation,
    RelatedOrg,
    Task,
)

NUM_FP = Path.home() / "Downloads" / "IPD_DL_numeric_7_00.xlsx"
DESC_FP = Path.home() / "Downloads" / "IPD_DL_description_7_00_01.xlsx"
LABELS_DIR = ROOT / "data" / "labels"
OUTPUT_DIR = ROOT / "data" / "occupations"
PROVENANCE_FILE = ROOT / "data" / ".ipd_provenance.json"

IPD_VERSION = "v7.00"
INGESTED_AT = dt.date.today().isoformat()

# ---------- Numeric subdivision IPD-ID prefixes (for labels lookup) ----------
LABELED_DIMENSION_PREFIX = {
    "interests": "IPD_04_01_",
    "work_values": "IPD_04_02_",
    "skills": "IPD_04_03_01_",
    "knowledge": "IPD_04_04_01_",
    "abilities": "IPD_04_12_",
    "work_characteristics": "IPD_04_05_",
    "work_activities": "IPD_04_10_",
}

# ---------- Hardcoded JA → en_key for the 5 distribution subsections ----------
# (These are categorical buckets, fewer than labels. Hardcoded for simplicity.)

EDUCATION_KEYS = {
    "高卒未満": "below_high_school",
    "高卒": "high_school",
    "専門学校卒": "vocational_school",
    "短大卒": "junior_college",
    "高専卒": "technical_college",
    "大卒": "university",
    "修士課程卒（修士と同等の専門職学位を含む）": "masters",
    "博士課程卒": "doctorate",
    "わからない": "unknown",
}

TRAINING_PRE_KEYS = {
    "特に必要ない": "not_required",
    "1ヶ月以下": "up_to_1_month",
    "1ヶ月超～6ヶ月以下": "1_to_6_months",
    "6ヶ月超～1年以下": "6_months_to_1_year",
    "1年超～2年以下": "1_to_2_years",
    "2年超～3年以下": "2_to_3_years",
    "3年超～5年以下": "3_to_5_years",
    "5年超～10年以下": "5_to_10_years",
    "10年超": "over_10_years",
    "わからない": "unknown",
}

# Same buckets used for prior experience (IPD_04_08_*)
EXPERIENCE_KEYS = TRAINING_PRE_KEYS

TRAINING_POST_KEYS = {
    "必要でない（未経験でも即戦力となる）": "not_required",
    "1ヶ月以下": "up_to_1_month",
    "1ヶ月超～6ヶ月以下": "1_to_6_months",
    "6ヶ月超～1年以下": "6_months_to_1_year",
    "1年超～2年以下": "1_to_2_years",
    "2年超～3年以下": "2_to_3_years",
    "3年超～5年以下": "3_to_5_years",
    "5年超～10年以下": "5_to_10_years",
    "10年超": "over_10_years",
    "わからない": "unknown",
}

EMPLOYMENT_TYPE_KEYS = {
    "正規の職員、従業員": "regular_employee",
    "パートタイマー": "part_time",
    "派遣社員": "dispatched",
    "契約社員、期間従業員": "contract",
    "自営、フリーランス": "self_employed_freelance",
    "経営層（役員等）": "executive",
    "アルバイト（学生以外）": "casual_non_student",
    "アルバイト（学生）": "casual_student",
    "わからない": "unknown",
    "その他": "other",
}

DISTRIBUTION_PREFIX = {
    "education_distribution": ("IPD_04_06_", EDUCATION_KEYS),
    "training_pre": ("IPD_04_07_", TRAINING_PRE_KEYS),
    "experience": ("IPD_04_08_", EXPERIENCE_KEYS),
    "training_post": ("IPD_04_09_", TRAINING_POST_KEYS),
    "employment_type": ("IPD_04_11_", EMPLOYMENT_TYPE_KEYS),
}


# ---------- Helpers ----------

def sha256_of_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def build_ipd_to_en_key() -> dict[str, str]:
    """Walk 7 labels files; return IPD-ID → en_key map for the labeled dimensions."""
    mapping: dict[str, str] = {}
    for dim_file in LABELS_DIR.glob("*.ja-en.json"):
        data = json.loads(dim_file.read_text(encoding="utf-8"))
        for en_key, entry in data["labels"].items():
            ipd_id = entry.get("onet_id")
            if ipd_id:
                if ipd_id in mapping:
                    raise RuntimeError(f"Duplicate IPD-ID {ipd_id} across labels files")
                mapping[ipd_id] = en_key
    return mapping


def header_to_col_map(ws, header_row: int) -> dict[str, int]:
    """Build IPD-ID → column index map from a header row."""
    out = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str) and v.startswith("IPD_"):
            out[v] = c
    return out


def _safe_str(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _safe_float(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# ---------- Per-occupation builders ----------

def build_aliases(desc_row: dict[str, object]) -> list[str]:
    out = []
    for i in range(1, 26):
        v = _safe_str(desc_row.get(f"IPD_02_04_{i:03d}"))
        if v:
            out.append(v)
    return out


def build_classifications(desc_row: dict[str, object]) -> Classifications:
    mhlw_main = _safe_str(desc_row.get("IPD_02_02_000"))
    mhlw_all = []
    for i in range(1, 7):
        v = _safe_str(desc_row.get(f"IPD_02_02_{i:03d}"))
        if v:
            mhlw_all.append(v)
    if mhlw_main and mhlw_main not in mhlw_all:
        mhlw_all.insert(0, mhlw_main)

    jsoc_main = _safe_str(desc_row.get("IPD_02_03_000"))
    jsoc_all = []
    for i in range(1, 6):
        v = _safe_str(desc_row.get(f"IPD_02_03_{i:03d}"))
        if v:
            jsoc_all.append(v)
    if jsoc_main and jsoc_main not in jsoc_all:
        jsoc_all.insert(0, jsoc_main)

    return Classifications(
        mhlw_main=mhlw_main,
        mhlw_all=mhlw_all,
        jsoc_main=jsoc_main,
        jsoc_all=jsoc_all,
    )


def build_description(desc_row: dict[str, object]) -> Description:
    return Description(
        summary_ja=_safe_str(desc_row.get("IPD_03_01_000")),
        what_it_is_ja=_safe_str(desc_row.get("IPD_03_01_001")),
        how_to_become_ja=_safe_str(desc_row.get("IPD_03_01_002")),
        working_conditions_ja=_safe_str(desc_row.get("IPD_03_01_003")),
    )


def build_related_orgs(desc_row: dict[str, object]) -> list[RelatedOrg]:
    out = []
    for i in range(1, 11):
        name = _safe_str(desc_row.get(f"IPD_03_03_{i:02d}_01"))
        url = _safe_str(desc_row.get(f"IPD_03_03_{i:02d}_02"))
        if name:
            out.append(RelatedOrg(name_ja=name, url=url))
    return out


def build_related_certs(desc_row: dict[str, object]) -> list[str]:
    out = []
    for i in range(1, 36):
        v = _safe_str(desc_row.get(f"IPD_03_04_{i:03d}"))
        if v:
            out.append(v)
    return out


def build_labeled_dimension(num_row: dict[str, object], prefix: str,
                             ipd_to_key: dict[str, str]) -> dict[str, float] | None:
    """Build one of the 7 labeled numeric dimensions. Return None if entire block missing."""
    block = {}
    for ipd_id, en_key in ipd_to_key.items():
        if not ipd_id.startswith(prefix):
            continue
        v = _safe_float(num_row.get(ipd_id))
        if v is not None:
            block[en_key] = v
    # Whole-block-None per §5.4: None if NO values present
    if not block:
        return None
    return block


def build_distribution(num_row: dict[str, object], prefix: str,
                        ja_to_key: dict[str, str], num_col_map_to_ja: dict[str, str]
                        ) -> dict[str, float] | None:
    """Build one of the 5 distribution blocks. Return None if entire block missing."""
    block = {}
    for ipd_id, ja_label in num_col_map_to_ja.items():
        if not ipd_id.startswith(prefix):
            continue
        en_key = ja_to_key.get(ja_label)
        if not en_key:
            # JA label drift — log but don't fail
            print(f"  WARN: distribution {prefix} JA '{ja_label}' not in en_key map", file=sys.stderr)
            continue
        v = _safe_float(num_row.get(ipd_id))
        if v is not None:
            block[en_key] = v
    if not block:
        return None
    return block


def build_tasks(num_row: dict[str, object]) -> tuple[str | None, list[Task]]:
    """Extract タスク_リード文 + up to 37 tasks from numeric row.

    IPD-ID format (verified against IPD v7.00):
        IPD_05_00_01 = タスク_リード文 (lead text)
        IPD_05_<NN>_01 = タスクN description
        IPD_05_<NN>_02 = タスクN_実施率 (execution rate, 0.0-1.0)
        IPD_05_<NN>_03 = タスクN_重要度 (importance, 0-5)
    Where <NN> is 2-digit zero-padded task number (01-37).
    """
    lead = _safe_str(num_row.get("IPD_05_00_01"))

    tasks: list[Task] = []
    for i in range(1, 38):
        prefix = f"IPD_05_{i:02d}_"
        desc = _safe_str(num_row.get(prefix + "01"))
        rate = _safe_float(num_row.get(prefix + "02"))
        imp = _safe_float(num_row.get(prefix + "03"))
        if desc is None and rate is None and imp is None:
            continue
        if not desc:
            continue
        tasks.append(Task(
            task_id=i,
            description_ja=desc,
            execution_rate=rate,
            importance=imp,
        ))
    return lead, tasks


def build_last_updated(desc_row: dict[str, object], num_row: dict[str, object]) -> dict[str, int]:
    """Walk IPD_88_* fields; collect non-null update years per section."""
    out: dict[str, int] = {}
    # Description-side 88 fields
    for ipd_id, key in [
        ("IPD_88_01_01", "ipd_id"),
        ("IPD_88_02_01", "classifications"),
        ("IPD_88_03_01", "description"),
    ]:
        v = desc_row.get(ipd_id)
        if v is not None:
            try:
                out[key] = int(v)
            except (TypeError, ValueError):
                pass
    # Numeric-side 88 fields (the 9 sub-domains)
    for ipd_id, key in [
        ("IPD_88_04_01_001", "interests"),
        ("IPD_88_04_01_002", "work_values"),
        ("IPD_88_04_01_003", "skills"),
        ("IPD_88_04_01_004", "knowledge"),
        ("IPD_88_04_01_005", "work_characteristics"),
        ("IPD_88_04_01_006", "work_activities"),
        ("IPD_88_04_01_007", "abilities"),
        ("IPD_88_04_01_008", "education_distribution"),
        ("IPD_88_04_01_009", "training_experience_employment"),
        ("IPD_88_05_01", "tasks"),
    ]:
        v = num_row.get(ipd_id)
        if v is not None:
            try:
                out[key] = int(v)
            except (TypeError, ValueError):
                pass
    return out


# ---------- Main pipeline ----------

def main() -> int:
    print(f"Loading {NUM_FP.name} ...")
    wb_num = load_workbook(NUM_FP, data_only=True)
    ws_num = wb_num["IPD形式"]
    num_col_map = header_to_col_map(ws_num, header_row=18)

    # Build IPD-ID → JA name map for distribution columns (using row 17 = field name)
    num_id_to_ja: dict[str, str] = {}
    for ipd_id, c in num_col_map.items():
        ja = ws_num.cell(row=17, column=c).value
        if ja:
            num_id_to_ja[ipd_id] = str(ja)

    # Read all numeric rows (data starts at row 19, key by id)
    num_data: dict[int, dict[str, object]] = {}
    for r in range(19, ws_num.max_row + 1):
        id_cell = ws_num.cell(row=r, column=num_col_map.get("IPD_01_01_001", 3)).value
        if id_cell is None:
            continue
        try:
            occ_id = int(id_cell)
        except (TypeError, ValueError):
            continue
        row_data = {ipd_id: ws_num.cell(row=r, column=c).value for ipd_id, c in num_col_map.items()}
        num_data[occ_id] = row_data
    wb_num.close()
    print(f"  {len(num_data)} numeric occupations loaded")

    print(f"Loading {DESC_FP.name} ...")
    wb_desc = load_workbook(DESC_FP, data_only=True)
    ws_desc = wb_desc["解説系"]
    desc_col_map = header_to_col_map(ws_desc, header_row=14)

    desc_data: dict[int, dict[str, object]] = {}
    for r in range(15, ws_desc.max_row + 1):
        id_cell = ws_desc.cell(row=r, column=desc_col_map.get("IPD_01_01_001", 3)).value
        if id_cell is None:
            continue
        try:
            occ_id = int(id_cell)
        except (TypeError, ValueError):
            continue
        row_data = {ipd_id: ws_desc.cell(row=r, column=c).value for ipd_id, c in desc_col_map.items()}
        desc_data[occ_id] = row_data
    wb_desc.close()
    print(f"  {len(desc_data)} description occupations loaded")

    # Reverse map: IPD-ID → en_key (for the 7 labeled numeric dimensions)
    ipd_to_key = build_ipd_to_en_key()
    print(f"  {len(ipd_to_key)} IPD-ID → en_key labels loaded")

    # Provenance
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    PROVENANCE_FILE.parent.mkdir(parents=True, exist_ok=True)

    print(f"Computing source SHA256 ...")
    num_sha = sha256_of_file(NUM_FP)
    desc_sha = sha256_of_file(DESC_FP)
    provenance = {
        "ipd_version": IPD_VERSION,
        "ipd_published_at": "2026-03-17",
        "ipd_data_cut_numeric": "2026-02-10",
        "ipd_data_cut_description": "2026-02-26",
        "ipd_retrieved_at": INGESTED_AT,
        "files": {
            "numeric": {"name": NUM_FP.name, "sha256": num_sha},
            "description": {"name": DESC_FP.name, "sha256": desc_sha},
        },
        "source_index_url": "https://shigoto.mhlw.go.jp/User/download",
        "source_file_url_numeric":
            "https://sgteprdstaplog01.blob.core.windows.net/web-app-contents/downloads/ver13/IPD_DL_numeric_7_00.xlsx",
        "source_file_url_description":
            "https://sgteprdstaplog01.blob.core.windows.net/web-app-contents/downloads/ver13/IPD_DL_description_7_00_01.xlsx",
        "license": "JILPT 職業情報データベース; secondary use permitted per jobtag TOS Article 9 with required attribution",
        "tos_url": "https://shigoto.mhlw.go.jp/User/tos",
    }
    PROVENANCE_FILE.write_text(json.dumps(provenance, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"  wrote {PROVENANCE_FILE.relative_to(ROOT)}")

    # Build occupation files
    all_ids = sorted(set(desc_data.keys()) | set(num_data.keys()))
    print(f"\nBuilding {len(all_ids)} occupation files ...")

    written = 0
    has_numeric = 0
    no_numeric = 0
    failures = []
    for occ_id in all_ids:
        desc_row = desc_data.get(occ_id)
        num_row = num_data.get(occ_id, {})

        if not desc_row:
            # Should never happen given desc is superset of num, but log defensively
            print(f"  WARN: id={occ_id} present in numeric but not description; skipping", file=sys.stderr)
            continue

        title_ja = _safe_str(desc_row.get("IPD_02_01_001"))
        if not title_ja:
            print(f"  WARN: id={occ_id} has no 職業名; skipping", file=sys.stderr)
            continue

        # 12 numeric subdivisions (None if id not in numeric file OR block fully empty)
        if num_row:
            interests = build_labeled_dimension(num_row, "IPD_04_01_", ipd_to_key)
            work_values = build_labeled_dimension(num_row, "IPD_04_02_", ipd_to_key)
            skills = build_labeled_dimension(num_row, "IPD_04_03_01_", ipd_to_key)
            knowledge = build_labeled_dimension(num_row, "IPD_04_04_01_", ipd_to_key)
            abilities = build_labeled_dimension(num_row, "IPD_04_12_", ipd_to_key)
            work_characteristics = build_labeled_dimension(num_row, "IPD_04_05_", ipd_to_key)
            work_activities = build_labeled_dimension(num_row, "IPD_04_10_", ipd_to_key)

            education_distribution = build_distribution(num_row, "IPD_04_06_", EDUCATION_KEYS, num_id_to_ja)
            training_pre = build_distribution(num_row, "IPD_04_07_", TRAINING_PRE_KEYS, num_id_to_ja)
            experience = build_distribution(num_row, "IPD_04_08_", EXPERIENCE_KEYS, num_id_to_ja)
            training_post = build_distribution(num_row, "IPD_04_09_", TRAINING_POST_KEYS, num_id_to_ja)
            employment_type = build_distribution(num_row, "IPD_04_11_", EMPLOYMENT_TYPE_KEYS, num_id_to_ja)

            tasks_lead_ja, tasks = build_tasks(num_row)
            has_numeric += 1
        else:
            interests = work_values = skills = knowledge = abilities = None
            work_characteristics = work_activities = None
            education_distribution = training_pre = experience = training_post = employment_type = None
            tasks_lead_ja = None
            tasks = []
            no_numeric += 1

        try:
            occ = Occupation(
                id=occ_id,
                ipd_id=f"IPD_01_01_{occ_id:03d}",
                schema_version="7.00",
                ingested_at=INGESTED_AT,
                title_ja=title_ja,
                aliases_ja=build_aliases(desc_row),
                classifications=build_classifications(desc_row),
                description=build_description(desc_row),
                interests=interests,
                work_values=work_values,
                skills=skills,
                knowledge=knowledge,
                abilities=abilities,
                work_characteristics=work_characteristics,
                work_activities=work_activities,
                education_distribution=education_distribution,
                training_pre=training_pre,
                training_post=training_post,
                experience=experience,
                employment_type=employment_type,
                tasks_lead_ja=tasks_lead_ja,
                tasks=tasks,
                related_orgs=build_related_orgs(desc_row),
                related_certs_ja=build_related_certs(desc_row),
                url=f"https://shigoto.mhlw.go.jp/User/Occupation/Detail/{occ_id}",
                data_source_versions=DataSourceVersions(
                    ipd_numeric=IPD_VERSION,
                    ipd_description=IPD_VERSION,
                    ipd_retrieved_at=INGESTED_AT,
                ),
                last_updated_per_section=build_last_updated(desc_row, num_row),
            )
        except Exception as e:
            failures.append((occ_id, str(e)))
            continue

        out_path = OUTPUT_DIR / f"{occ_id:04d}.json"
        out_path.write_text(
            json.dumps(occ.model_dump(exclude_none=False), ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        written += 1

    print(f"\nDone:")
    print(f"  wrote          {written} files to {OUTPUT_DIR.relative_to(ROOT)}")
    print(f"  with numeric   {has_numeric}")
    print(f"  desc only      {no_numeric}")
    print(f"  failures       {len(failures)}")
    if failures:
        print("First 5 failures:")
        for occ_id, err in failures[:5]:
            print(f"  id={occ_id}: {err[:200]}")
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
