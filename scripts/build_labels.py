"""Generate data/labels/*.ja-en.json from IPD 細目 sheet.

One-shot generator per DATA_ARCHITECTURE.md §2.5 / D-011.

Reads:
    ~/Downloads/IPD_DL_numeric_7_00.xlsx (細目 sheet) for JA names + IPD-IDs
Writes:
    data/labels/{interests,work_values,skills,knowledge,abilities,
                 work_characteristics,work_activities}.ja-en.json

REAL counts per dimension (verified against IPD v7.00):
    interests=6, work_values=11, skills=39, knowledge=33, abilities=35,
    work_characteristics=39, work_activities=41   →  total 204
(Doc v1.0.5 §2.1/§2.5 had skills=78 / knowledge=66 — those numbers included
 _無関係フラグ flag fields. Corrected in v1.0.6.)

EN names: best-effort O*NET-aligned + literal where JILPT diverges.
Marked status='draft v0.1' — verify before production use.

Key naming: snake_case English, derived from EN label.

Usage:
    uv run python scripts/build_labels.py
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "data"))
from schema.labels import LabelEntry, LabelsFile  # noqa: E402

NUM_FP = Path.home() / "Downloads" / "IPD_DL_numeric_7_00.xlsx"
OUTPUT_DIR = ROOT / "data" / "labels"

# Strict IPD-ID prefix per dimension (excludes _無関係フラグ subsections like _02_)
DIMENSION_PREFIX = {
    "interests": "IPD_04_01_",
    "work_values": "IPD_04_02_",
    "skills": "IPD_04_03_01_",
    "knowledge": "IPD_04_04_01_",
    "abilities": "IPD_04_12_",
    "work_characteristics": "IPD_04_05_",
    "work_activities": "IPD_04_10_",
}

EXPECTED_COUNTS = {
    "interests": 6, "work_values": 11, "skills": 39, "knowledge": 33,
    "abilities": 35, "work_characteristics": 39, "work_activities": 41,
}

# ---------- JA → (en_key, en_label) mapping ----------
# Total 204 entries across 7 dimensions. EN: best-effort O*NET-aligned.

INTERESTS: dict[str, tuple[str, str]] = {
    "現実的": ("realistic", "Realistic"),
    "研究的": ("investigative", "Investigative"),
    "芸術的": ("artistic", "Artistic"),
    "社会的": ("social", "Social"),
    "企業的": ("enterprising", "Enterprising"),
    "慣習的": ("conventional", "Conventional"),
}

WORK_VALUES: dict[str, tuple[str, str]] = {
    "達成感": ("achievement", "Achievement"),
    "自律性": ("autonomy", "Autonomy"),
    "専門性": ("expertise", "Expertise"),
    "自己成長": ("self_growth", "Self-Growth"),
    "社会的認知・地位": ("social_recognition", "Social Recognition / Status"),
    "奉仕・社会貢献": ("service_contribution", "Service & Social Contribution"),
    "良好な対人関係": ("good_relationships", "Good Interpersonal Relationships"),
    "労働安全衛生": ("workplace_safety", "Workplace Safety & Health"),
    "私生活との両立": ("work_life_balance", "Work-Life Balance"),
    "雇用や生活の安定性": ("stability", "Employment & Life Stability"),
    "報酬や収入": ("compensation", "Compensation & Income"),
}

SKILLS: dict[str, tuple[str, str]] = {
    "読解力": ("reading_comprehension", "Reading Comprehension"),
    "傾聴力": ("active_listening", "Active Listening"),
    "文章力": ("writing", "Writing"),
    "説明力": ("speaking", "Speaking"),
    "外国語を読む": ("foreign_language_reading", "Foreign Language Reading"),
    "外国語を聞く": ("foreign_language_listening", "Foreign Language Listening"),
    "外国語で書く": ("foreign_language_writing", "Foreign Language Writing"),
    "外国語で話す": ("foreign_language_speaking", "Foreign Language Speaking"),
    "数学的素養": ("mathematics", "Mathematics"),
    "科学的素養": ("science", "Science"),
    "論理と推論（批判的思考）": ("critical_thinking", "Critical Thinking"),
    "新しい情報の応用力": ("active_learning", "Active Learning"),
    "学習方法の選択・実践": ("learning_strategies", "Learning Strategies"),
    "継続的観察と評価": ("monitoring", "Monitoring"),
    "他者の反応の理解": ("social_perceptiveness", "Social Perceptiveness"),
    "他者との調整": ("coordination", "Coordination"),
    "説得": ("persuasion", "Persuasion"),
    "交渉": ("negotiation", "Negotiation"),
    "指導": ("instructing", "Instructing"),
    "対人援助サービス": ("service_orientation", "Service Orientation"),
    "複雑な問題解決": ("complex_problem_solving", "Complex Problem Solving"),
    "要件分析（仕様作成）": ("operations_analysis", "Operations Analysis"),
    "カスタマイズと開発": ("technology_design", "Technology Design"),
    "道具、機器、設備の選択": ("equipment_selection", "Equipment Selection"),
    "設置と設定": ("installation", "Installation"),
    "プログラミング": ("programming", "Programming"),
    "計器監視": ("operations_monitoring", "Operations Monitoring"),
    "操作と制御": ("operation_and_control", "Operation and Control"),
    "保守点検": ("equipment_maintenance", "Equipment Maintenance"),
    "故障等の原因特定": ("troubleshooting", "Troubleshooting"),
    "修理": ("repairing", "Repairing"),
    "クオリティチェック": ("quality_control_analysis", "Quality Control Analysis"),
    "合理的な意思決定": ("judgment_decision_making", "Judgment & Decision Making"),
    "企業・組織の活動の分析": ("systems_analysis", "Systems Analysis"),
    "企業・組織の活動の評価": ("systems_evaluation", "Systems Evaluation"),
    "時間管理": ("time_management", "Time Management"),
    "資金管理": ("financial_resources_management", "Management of Financial Resources"),
    "資材管理": ("material_resources_management", "Management of Material Resources"),
    "人材管理": ("personnel_resources_management", "Management of Personnel Resources"),
}

KNOWLEDGE: dict[str, tuple[str, str]] = {
    "ビジネスと経営": ("business_management", "Business & Management"),
    "事務処理": ("clerical", "Clerical"),
    "経済学・会計学": ("economics_accounting", "Economics & Accounting"),
    "販売・マーケティング": ("sales_marketing", "Sales & Marketing"),
    "顧客サービス・対人サービス": ("customer_personal_service", "Customer & Personal Service"),
    "人事労務管理": ("personnel_hr", "Personnel & Human Resources"),
    "輸送": ("transportation", "Transportation"),
    "生産・加工": ("production_processing", "Production & Processing"),
    "農業・畜産業": ("agriculture_livestock", "Agriculture & Livestock"),
    "工学": ("engineering", "Engineering"),
    "コンピュータと電子工学": ("computers_electronics", "Computers & Electronics"),
    "設計": ("design", "Design"),
    "建築・建設": ("building_construction", "Building & Construction"),
    "機械": ("mechanical", "Mechanical"),
    "数学": ("mathematics_kn", "Mathematics"),
    "物理学": ("physics", "Physics"),
    "化学": ("chemistry", "Chemistry"),
    "生物学": ("biology", "Biology"),
    "心理学": ("psychology", "Psychology"),
    "社会学": ("sociology", "Sociology"),
    "地理学": ("geography", "Geography"),
    "医学・歯学": ("medicine_dentistry", "Medicine & Dentistry"),
    "セラピーとカウンセリング": ("therapy_counseling", "Therapy & Counseling"),
    "教育訓練": ("education_training_kn", "Education & Training"),
    "日本語の語彙・文法": ("japanese_language", "Japanese Language (Vocabulary & Grammar)"),
    "外国語の語彙・文法": ("foreign_language_kn", "Foreign Language (Vocabulary & Grammar)"),
    "芸術": ("fine_arts", "Fine Arts"),
    "歴史学・考古学": ("history_archaeology", "History & Archaeology"),
    "哲学・宗教学": ("philosophy_theology", "Philosophy & Theology"),
    "公衆安全・危機管理": ("public_safety_security", "Public Safety & Security"),
    "法律学、政治学": ("law_government", "Law & Government / Political Science"),
    "通信技術": ("telecommunications", "Telecommunications Technology"),
    "コミュニケーションとメディア": ("communications_media", "Communications & Media"),
}

ABILITIES: dict[str, tuple[str, str]] = {
    "発話理解": ("oral_comprehension", "Oral Comprehension"),
    "記述理解": ("written_comprehension", "Written Comprehension"),
    "発話表現": ("oral_expression", "Oral Expression"),
    "記述表現": ("written_expression", "Written Expression"),
    "アイデアや代案を数多く生み出す力": ("fluency_of_ideas", "Fluency of Ideas"),
    "独創性": ("originality", "Originality"),
    "トラブルの察知": ("problem_sensitivity", "Problem Sensitivity"),
    "演繹的推論": ("deductive_reasoning", "Deductive Reasoning"),
    "帰納的推論": ("inductive_reasoning", "Inductive Reasoning"),
    "法則に基づいた情報の並べ替え": ("information_ordering", "Information Ordering"),
    "カテゴライズ": ("category_flexibility", "Category Flexibility"),
    "数学的推論": ("mathematical_reasoning", "Mathematical Reasoning"),
    "演算力": ("number_facility", "Number Facility"),
    "記憶力": ("memorization", "Memorization"),
    "知覚速度": ("perceptual_speed", "Perceptual Speed"),
    "自他の位置関係の把握": ("spatial_orientation", "Spatial Orientation"),
    "モノの見え方に関する想像力": ("visualization", "Visualization"),
    "選択的注意（集中する力）": ("selective_attention", "Selective Attention"),
    "マルチタスク": ("time_sharing", "Time Sharing / Multitasking"),
    "腕と手の安定": ("arm_hand_steadiness", "Arm-Hand Steadiness"),
    "手腕の器用さ": ("manual_dexterity", "Manual Dexterity"),
    "指先の器用さ": ("finger_dexterity", "Finger Dexterity"),
    "一瞬で素早く反応する力": ("reaction_time", "Reaction Time"),
    "手首と指の動作速度": ("wrist_finger_speed", "Wrist-Finger Speed"),
    "腕や脚の動作速度": ("speed_of_limb_movement", "Speed of Limb Movement"),
    "筋力": ("static_strength", "Static Strength"),
    "持久力（スタミナ）": ("stamina", "Stamina"),
    "平衡感覚": ("gross_body_equilibrium", "Gross Body Equilibrium"),
    "近接視力": ("near_vision", "Near Vision"),
    "遠隔視力": ("far_vision", "Far Vision"),
    "色の違いを見分ける力": ("visual_color_discrimination", "Visual Color Discrimination"),
    "奥行きの知覚（遠近感覚、深視力）": ("depth_perception", "Depth Perception"),
    "聴覚の感度": ("hearing_sensitivity", "Hearing Sensitivity"),
    "聴覚的注意（特定の音を聞き分ける力）": ("auditory_attention", "Auditory Attention"),
    "発話明瞭性": ("speech_clarity", "Speech Clarity"),
}

WORK_CHARACTERISTICS: dict[str, tuple[str, str]] = {
    "他者とのかかわり": ("contact_with_others", "Contact with Others"),
    "対面での議論": ("face_to_face_discussions", "Face-to-Face Discussions"),
    "電話での会話": ("telephone_communication", "Telephone Communication"),
    "ビジネスレターやメモの作成": ("business_writing", "Writing Business Letters & Memos"),
    "仕事上での他者との対立": ("conflict_situations", "Workplace Conflict Situations"),
    "時間的切迫": ("time_pressure", "Time Pressure"),
    "グループやチームでの仕事": ("teamwork", "Working in a Group or Team"),
    "外部の顧客等との接触": ("dealing_with_external_customers", "Dealing with External Customers"),
    "他者と調整し、リードする": ("coordinate_lead_others", "Coordinating & Leading Others"),
    "厳密さ、正確さ": ("exactness_accuracy", "Exactness & Accuracy"),
    "同一作業の反復": ("repetitive_tasks", "Repetitive Tasks"),
    "機器等の速度に応じた作業": ("pace_determined_by_machine", "Pace Determined by Machine Speed"),
    "結果・成果への責任": ("responsibility_for_outcomes", "Responsibility for Outcomes"),
    "空調のきいた屋内作業": ("indoor_climate_controlled", "Indoor Climate-Controlled Environment"),
    "空調のきいていない屋内作業": ("indoor_non_climate_controlled", "Indoor Non-Climate-Controlled Environment"),
    "屋外作業": ("outdoor_work", "Outdoor Work"),
    "座り作業": ("sitting", "Sitting"),
    "立ち作業": ("standing", "Standing"),
    "反復作業": ("repetition_of_activities", "Repetition of Activities"),
    "ミスの影響度": ("consequence_of_error", "Consequence of Error"),
    "意思決定の自由": ("freedom_to_make_decisions", "Freedom to Make Decisions"),
    "優先順位や目標の自己設定": ("set_own_goals", "Setting Own Goals & Priorities"),
    "規則的（ルーチンやスケジュールが決まっている）": ("regular_schedule", "Regular Schedule (Routine)"),
    "不規則（天候、生産需要、契約期間などで変わる）": ("irregular_schedule", "Irregular Schedule"),
    "季節的（一年のうちの一定の時期だけ）": ("seasonal_schedule", "Seasonal Schedule"),
    "電子メール": ("email_communication", "Email Communication"),
    "窮屈な仕事の場所、居心地が悪い姿勢": ("cramped_uncomfortable", "Cramped Work Space / Uncomfortable Positions"),
    "病気、感染症のリスク": ("disease_infection_risk", "Risk of Disease & Infection"),
    "軽度の火傷、切り傷、噛まれ傷、刺し傷": ("minor_injury_risk", "Risk of Minor Burns, Cuts, Bites, Stings"),
    "一般的な保護・安全装備の着用": ("standard_protective_equipment", "Wearing Standard Protective Equipment"),
    "特殊な保護・安全装備の着用": ("specialized_protective_equipment", "Wearing Specialized Protective Equipment"),
    "暴力的な人々への対応": ("dealing_with_violent_people", "Dealing with Violent People"),
    "歩行、走行": ("walking_running", "Walking & Running"),
    "モノ、道具、制御装置を扱う手作業": ("handling_objects_tools", "Manual Handling of Objects, Tools & Controls"),
    "他者との身体的近接": ("physical_proximity", "Physical Proximity to Others"),
    "機械やコンピュータによる仕事の自動化": ("work_automation", "Work Automation by Machines / Computers"),
    "他者の健康・安全への責任": ("responsibility_for_safety", "Responsibility for Others' Health & Safety"),
    "意思決定が他者や企業に及ぼす影響力": ("impact_of_decisions", "Impact of Decisions on Others / Organization"),
    "競争水準": ("level_of_competition", "Level of Competition"),
}

WORK_ACTIVITIES: dict[str, tuple[str, str]] = {
    "情報を取得する": ("getting_information", "Getting Information"),
    "継続的に状況を把握する": ("monitor_processes", "Monitor Processes / Materials / Surroundings"),
    "情報の整理と検知を行う": ("identifying_objects", "Identifying Objects, Actions & Events"),
    "設備、構造物、材料を検査する": ("inspecting_equipment", "Inspecting Equipment, Structures & Materials"),
    "数値の算出・推計を行う": ("estimating_quantifiable", "Estimating Quantifiable Characteristics"),
    "クオリティを判断する": ("judging_quality", "Judging the Qualities of Things"),
    "法律や規定、基準を適用する": ("evaluating_compliance", "Evaluating Information for Compliance"),
    "情報やデータを処理する": ("processing_information", "Processing Information"),
    "情報やデータを分析する": ("analyzing_data", "Analyzing Data or Information"),
    "意思決定と問題解決を行う": ("making_decisions", "Making Decisions & Solving Problems"),
    "創造的に考える": ("thinking_creatively", "Thinking Creatively"),
    "仕事に関連する知識を更新し、活用する": ("updating_knowledge", "Updating & Using Relevant Knowledge"),
    "目標と戦略を策定する": ("developing_objectives", "Developing Objectives & Strategies"),
    "スケジュールを作成する": ("scheduling_work", "Scheduling Work & Activities"),
    "仕事を整理、計画する、優先順序を決める": ("organizing_planning", "Organizing, Planning & Prioritizing Work"),
    "全身を使って身体的な活動を行う": ("physical_activities", "Performing General Physical Activities"),
    "手と腕を使って物を取り扱い動かす": ("handling_objects", "Handling & Moving Objects"),
    "機械、および機械製造のプロセスをコントロールする": ("controlling_machines", "Controlling Machines & Processes"),
    "乗り物を運転・操縦する": ("operating_vehicles", "Operating Vehicles & Mechanized Devices"),
    "コンピュータを用いて作業を行う": ("computer_work", "Interacting with Computers"),
    "装置、部品、機器の図面を作成する、配列や仕様を設定する": ("drafting_specifications", "Drafting, Laying Out & Specifying Equipment"),
    "機械装置の修理と保守を行う": ("repairing_mechanical", "Repairing & Maintaining Mechanical Equipment"),
    "電子機器の修理と保守を行う": ("repairing_electronic", "Repairing & Maintaining Electronic Equipment"),
    "情報の文書化と記録を行う": ("documenting_information", "Documenting / Recording Information"),
    "情報の意味を他者に説明する": ("interpreting_information", "Interpreting Information for Others"),
    "上司、同僚、部下とコミュニケーションを取る": ("communicating_inside", "Communicating with Supervisors / Peers / Subordinates"),
    "組織外の人々とコミュニケーションを取る": ("communicating_outside", "Communicating with People Outside the Organization"),
    "人間関係を構築し、維持する": ("building_relationships", "Establishing & Maintaining Relationships"),
    "他者に対する支援とケアを行う": ("assisting_caring", "Assisting & Caring for Others"),
    "他者に対して売り込む、または他者の思考・行動が変容するよう働きかける": ("selling_influencing", "Selling or Influencing Others"),
    "対立を解消させる、他者と交渉する": ("resolving_conflicts", "Resolving Conflicts & Negotiating"),
    "公共の場で一般の人々のために働いたり、直接応対する": ("performing_for_public", "Performing for or Working Directly with the Public"),
    "メンバーの仕事量や活動内容を調整する": ("coordinating_others", "Coordinating Work & Activities of Others"),
    "チームを構築する": ("developing_teams", "Developing & Building Teams"),
    "他者の訓練と教育を行う": ("training_teaching", "Training & Teaching Others"),
    "部下への指導、指示、動機づけを行う": ("guiding_subordinates", "Guiding, Directing & Motivating Subordinates"),
    "他者をコーチし、能力開発を行う": ("coaching_developing", "Coaching & Developing Others"),
    "コンサルティングと他者へのアドバイスを行う": ("providing_consultation", "Providing Consultation & Advice"),
    "管理業務を遂行する": ("performing_administrative", "Performing Administrative Activities"),
    "組織の人事管理を行う": ("staffing_organizational", "Staffing Organizational Units"),
    "資源、資材、財源の監視と管理を行う": ("monitoring_resources", "Monitoring & Controlling Resources"),
}

ALL_DIMENSIONS = {
    "interests": INTERESTS,
    "work_values": WORK_VALUES,
    "skills": SKILLS,
    "knowledge": KNOWLEDGE,
    "abilities": ABILITIES,
    "work_characteristics": WORK_CHARACTERISTICS,
    "work_activities": WORK_ACTIVITIES,
}


def main() -> int:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(NUM_FP, data_only=True)
    ws = wb["インプットデータ細目"]

    # Strict prefix-based extraction (skip _02_ flag subsections)
    extracted: dict[str, list[tuple[str, str]]] = {dim: [] for dim in DIMENSION_PREFIX}
    for r in range(4, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, 13)]
        _, dai, chu, sho, ver, ipd_id, name, dtype, rng, desc, memo, _ = row
        if not ipd_id or dtype != "数値":
            continue
        for dim_key, prefix in DIMENSION_PREFIX.items():
            if str(ipd_id).startswith(prefix):
                extracted[dim_key].append((ipd_id, name))
                break
    wb.close()

    # Validate counts and presence in our hand-curated mapping
    errors: list[str] = []
    for dim, items in extracted.items():
        expected = EXPECTED_COUNTS[dim]
        if len(items) != expected:
            errors.append(f"  {dim}: expected {expected}, got {len(items)}")
        mapping = ALL_DIMENSIONS[dim]
        for ipd_id, ja in items:
            if ja not in mapping:
                errors.append(f"  {dim}: missing EN translation for JA '{ja}' (IPD-ID {ipd_id})")
    if errors:
        print("ERRORS detected:", file=sys.stderr)
        for e in errors:
            print(e, file=sys.stderr)
        return 1

    # Write 7 labels files
    written = 0
    total_labels = 0
    for dim, items in extracted.items():
        mapping = ALL_DIMENSIONS[dim]
        labels: dict[str, LabelEntry] = {}
        for ipd_id, ja in items:
            en_key, en_label = mapping[ja]
            if en_key in labels:
                print(f"ERROR: duplicate key '{en_key}' in {dim}", file=sys.stderr)
                return 1
            labels[en_key] = LabelEntry(ja=ja, en=en_label, onet_id=ipd_id, note=None)

        labels_file = LabelsFile(
            dimension=dim,
            source="IPD 細目 v7.00 (JA) + Claude-curated O*NET-aligned EN (draft v0.1)",
            license="JA labels: see jobtag TOS Article 9; EN labels: original work, public domain",
            count=len(labels),
            labels=labels,
        )

        out_path = OUTPUT_DIR / f"{dim}.ja-en.json"
        out_path.write_text(
            json.dumps(labels_file.model_dump(exclude_none=True), ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        print(f"Wrote {out_path.relative_to(ROOT)}  ({len(labels)} labels)")
        written += 1
        total_labels += len(labels)

    print(f"\nDone. {written} labels files / {total_labels} total labels written.")
    print("Status: draft v0.1 — EN names need O*NET cross-reference review before production.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
